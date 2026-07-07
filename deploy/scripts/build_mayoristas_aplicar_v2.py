#!/usr/bin/env python3
"""Genera script SQL corregido para aplicar descuentos mayoristas."""
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

STG = 'stg_mayoristas_descuento'


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip())


def norm_key(s):
    s = norm(s).lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^a-z0-9 ]', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def split_name(full):
    parts = norm(full).split()
    if not parts:
        return '', '', ''
    if len(parts) == 1:
        return parts[0], '', ''
    if len(parts) == 2:
        return parts[0], parts[1], ''
    return ' '.join(parts[:-2]), parts[-2], parts[-1]


def parse_discount(val):
    if val is None or str(val).strip() in ('', '/', '-'):
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    pct = round(f * 100, 2) if 0 < f <= 1 else round(f, 2)
    return pct if 0 <= pct <= 100 else None


def esc(s):
    return s.replace('\\', '\\\\').replace("'", "''")


def norm_sql_expr(expr):
    return f"""LOWER(TRIM(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
    {expr},
    CHAR(225),'a'),CHAR(233),'e'),CHAR(237),'i'),CHAR(243),'o'),CHAR(250),'u')))"""


def load_rows(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        if sheet == 'Hoja1':
            continue
        for r in wb[sheet].iter_rows(min_row=1, values_only=True):
            if not r:
                continue
            name = norm(r[0]) if len(r) > 0 else ''
            if not name or name in ('/', '-'):
                continue
            disc_raw = r[3] if sheet == 'A' and len(r) >= 4 and r[3] is not None else (r[1] if len(r) >= 2 else None)
            pct = parse_discount(disc_raw)
            if pct is None:
                continue
            p_n, p_a1, p_a2 = split_name(name)
            rows.append({
                'sheet': sheet,
                'nombre_completo': name,
                'descuento_pct': pct,
                'nombre_norm': norm_key(name),
                'p_nombre_norm': norm_key(p_n),
                'p_ap1_norm': norm_key(p_a1),
                'p_ap2_norm': norm_key(p_a2),
            })
    return rows


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads' / 'MAYORISTAS-1.xlsx'
    out = Path(__file__).resolve().parents[2] / 'sql' / '2026_07_06_mayoristas_descuentos_aplicar_v2.sql'
    rows = load_rows(xlsx)

    full_norm = norm_sql_expr("CONCAT_WS(' ', u.nombre, u.primer_apellido, u.segundo_apellido)")
    n_norm = norm_sql_expr('u.nombre')
    a1_norm = norm_sql_expr('u.primer_apellido')
    a2_norm = norm_sql_expr("IFNULL(u.segundo_apellido, '')")

    match_sql = f"""(
    {full_norm} = m.nombre_norm
    OR (
        {n_norm} = m.p_nombre_norm
        AND {a1_norm} = m.p_ap1_norm
        AND (m.p_ap2_norm = '' OR {a2_norm} = m.p_ap2_norm)
    )
)"""

    lines = [
        '-- Descuentos mayoristas v2 (tabla permanente + match doble)',
        f'-- Registros Excel: {len(rows)}',
        '-- IMPORTANTE: ejecutar TODO el archivo de una vez (Ctrl+A -> Run)',
        '-- ROW_COUNT() en DataGrip devuelve -1; usar los SELECT de verificacion al final',
        '',
        f'DROP TABLE IF EXISTS {STG};',
        f'CREATE TABLE {STG} (',
        '  id INT AUTO_INCREMENT PRIMARY KEY,',
        '  hoja VARCHAR(10) NOT NULL,',
        '  nombre_completo VARCHAR(255) NOT NULL,',
        '  descuento_porcentaje DECIMAL(5,2) NOT NULL,',
        '  nombre_norm VARCHAR(255) NOT NULL,',
        '  p_nombre_norm VARCHAR(120) NOT NULL,',
        '  p_ap1_norm VARCHAR(80) NOT NULL,',
        '  p_ap2_norm VARCHAR(80) NOT NULL,',
        '  INDEX idx_nombre_norm (nombre_norm),',
        '  INDEX idx_partes (p_nombre_norm, p_ap1_norm, p_ap2_norm)',
        ') ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;',
        '',
    ]

    batch = [
        "('{hs}', '{nc}', {dp:.2f}, '{nn}', '{pn}', '{pa1}', '{pa2}')".format(
            hs=esc(r['sheet']),
            nc=esc(r['nombre_completo']),
            dp=r['descuento_pct'],
            nn=esc(r['nombre_norm']),
            pn=esc(r['p_nombre_norm']),
            pa1=esc(r['p_ap1_norm']),
            pa2=esc(r['p_ap2_norm']),
        )
        for r in rows
    ]
    cols = f'INSERT INTO {STG} (hoja, nombre_completo, descuento_porcentaje, nombre_norm, p_nombre_norm, p_ap1_norm, p_ap2_norm) VALUES'
    for i in range(0, len(batch), 60):
        lines.append(cols)
        lines.append(',\n'.join(batch[i:i + 60]) + ';')
        lines.append('')

    lines.extend([
        '-- Cuántos registros cargó el Excel',
        f'SELECT COUNT(*) AS registros_excel FROM {STG};',
        '',
        '-- Cuántos clientes coinciden ANTES del update',
        f'''SELECT COUNT(DISTINCT c.id_cliente) AS clientes_a_actualizar
FROM {STG} m
INNER JOIN usuarios u ON u.activo = 1
INNER JOIN clientes c ON c.id_usuario_FK = u.id_usuario AND c.activo = 1
WHERE {match_sql}
  AND (c.descuento_porcentaje IS NULL OR CAST(c.descuento_porcentaje AS DECIMAL(5,2)) <> m.descuento_porcentaje);''',
        '',
        '-- UPDATE (ejecutar en la misma sesion que los INSERT)',
        f'''UPDATE clientes c
INNER JOIN usuarios u ON u.id_usuario = c.id_usuario_FK AND u.activo = 1
INNER JOIN {STG} m ON {match_sql}
SET c.descuento_porcentaje = m.descuento_porcentaje
WHERE c.activo = 1
  AND (c.descuento_porcentaje IS NULL OR CAST(c.descuento_porcentaje AS DECIMAL(5,2)) <> m.descuento_porcentaje);''',
        '',
        '-- Verificacion: clientes con descuento 40 o 50 despues del update',
        '''SELECT CAST(descuento_porcentaje AS DECIMAL(5,2)) AS descuento, COUNT(*) AS total
FROM clientes
WHERE activo = 1 AND descuento_porcentaje IN (40, 50)
GROUP BY CAST(descuento_porcentaje AS DECIMAL(5,2));''',
        '',
        '-- Excel sin match en BD (revisar manualmente)',
        f'''SELECT m.hoja, m.nombre_completo, m.descuento_porcentaje
FROM {STG} m
WHERE NOT EXISTS (
    SELECT 1 FROM clientes c
    INNER JOIN usuarios u ON u.id_usuario = c.id_usuario_FK AND u.activo = 1
    WHERE c.activo = 1 AND {match_sql}
)
ORDER BY m.nombre_completo;''',
        '',
        f'-- Opcional: DROP TABLE {STG};  -- descomentar cuando confirmes',
    ])

    out.write_text('\n'.join(lines), encoding='utf-8')
    print(f'OK {len(rows)} filas -> {out}')


if __name__ == '__main__':
    main()
