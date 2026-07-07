#!/usr/bin/env python3
"""Parse MAYORISTAS Excel and emit SQL for DataGrip."""
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl


def norm(s):
    if s is None:
        return ''
    s = str(s).strip()
    return re.sub(r'\s+', ' ', s)


def norm_key(s):
    s = norm(s).lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^a-z0-9 ]', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def parse_discount(val):
    if val is None or str(val).strip() in ('', '/', '-'):
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    pct = round(f * 100, 2) if 0 < f <= 1 else round(f, 2)
    return pct if 0 <= pct <= 100 else None


def esc(s):
    return s.replace('\\', '\\\\').replace("'", "''")


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads' / 'MAYORISTAS-1.xlsx'
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parents[2] / 'sql' / '2026_07_06_mayoristas_descuentos_import.sql'

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    rows = []
    skipped = []

    for sheet in wb.sheetnames:
        if sheet == 'Hoja1':
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=1, values_only=True):
            if not r:
                continue
            name = norm(r[0]) if len(r) > 0 else ''
            if not name or name in ('/', '-'):
                skipped.append((sheet, name))
                continue
            disc_raw = None
            if sheet == 'A' and len(r) >= 4 and r[3] is not None:
                disc_raw = r[3]
            elif len(r) >= 2 and r[1] is not None:
                disc_raw = r[1]
            pct = parse_discount(disc_raw)
            if pct is None:
                skipped.append((sheet, name))
                continue
            rows.append({
                'sheet': sheet,
                'nombre_completo': name,
                'descuento_pct': pct,
                'key': norm_key(name),
            })

    by_key = {}
    dups = []
    for row in rows:
        if row['key'] in by_key:
            dups.append((by_key[row['key']], row))
        else:
            by_key[row['key']] = row

    disc_counts = Counter(r['descuento_pct'] for r in rows)
    print(f'TOTAL: {len(rows)} | OMITIDOS: {len(skipped)} | DUPLICADOS: {len(dups)}')
    print('DESCUENTOS:', dict(sorted(disc_counts.items())))
    for r in rows:
        if r['descuento_pct'] not in (40.0, 50.0):
            print(f'ATIPICO: {r["nombre_completo"]} -> {r["descuento_pct"]}%')

    norm_sql = """LOWER(TRIM(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
    CONCAT_WS(' ', u.nombre, u.primer_apellido, u.segundo_apellido),
    'á','a'),'é','e'),'í','i'),'ó','o'),'ú','u')))"""

    lines = [
        '-- Generado desde MAYORISTAS-1.xlsx',
        f'-- Total registros: {len(rows)}',
        '-- Excel: 0.4 = 40%, 0.5 = 50% -> clientes.descuento_porcentaje',
        '-- 1) Ejecuta INSERTs + SELECT preview',
        '-- 2) Revisa NO_ENCONTRADO y duplicados',
        '-- 3) Ejecuta UPDATE masivo si todo cuadra',
        '',
        'DROP TEMPORARY TABLE IF EXISTS tmp_mayoristas_excel;',
        'CREATE TEMPORARY TABLE tmp_mayoristas_excel (',
        '  id INT AUTO_INCREMENT PRIMARY KEY,',
        '  hoja VARCHAR(10) NOT NULL,',
        '  nombre_completo VARCHAR(255) NOT NULL,',
        '  descuento_porcentaje DECIMAL(5,2) NOT NULL,',
        '  nombre_norm VARCHAR(255) NOT NULL,',
        '  INDEX idx_nombre_norm (nombre_norm)',
        ');',
        '',
    ]

    batch = [
        f"('{esc(r['sheet'])}', '{esc(r['nombre_completo'])}', {r['descuento_pct']:.2f}, '{esc(r['key'])}')"
        for r in rows
    ]
    for i in range(0, len(batch), 80):
        part = batch[i:i + 80]
        lines.append('INSERT INTO tmp_mayoristas_excel (hoja, nombre_completo, descuento_porcentaje, nombre_norm) VALUES')
        lines.append(',\n'.join(part) + ';')
        lines.append('')

    lines.extend([
        '-- PREVIEW: cruce Excel vs BD',
        f'''SELECT
    m.hoja,
    m.nombre_completo AS excel_nombre,
    m.descuento_porcentaje AS excel_descuento,
    c.id_cliente,
    TRIM(CONCAT_WS(' ', u.nombre, u.primer_apellido, u.segundo_apellido)) AS bd_nombre,
    c.descuento_porcentaje AS bd_descuento_actual,
    CASE
        WHEN c.id_cliente IS NULL THEN 'NO_ENCONTRADO'
        WHEN CAST(c.descuento_porcentaje AS DECIMAL(5,2)) = m.descuento_porcentaje THEN 'OK_IGUAL'
        ELSE 'ACTUALIZAR'
    END AS accion
FROM tmp_mayoristas_excel m
LEFT JOIN usuarios u ON u.activo = 1
LEFT JOIN clientes c ON c.id_usuario_FK = u.id_usuario AND c.activo = 1
    AND {norm_sql} = m.nombre_norm
ORDER BY accion, m.nombre_completo;''',
        '',
        '-- RESUMEN',
        f'''SELECT accion, COUNT(*) AS total FROM (
    SELECT CASE
        WHEN c.id_cliente IS NULL THEN 'NO_ENCONTRADO'
        WHEN CAST(c.descuento_porcentaje AS DECIMAL(5,2)) = m.descuento_porcentaje THEN 'OK_IGUAL'
        ELSE 'ACTUALIZAR'
    END AS accion
    FROM tmp_mayoristas_excel m
    LEFT JOIN usuarios u ON u.activo = 1
    LEFT JOIN clientes c ON c.id_usuario_FK = u.id_usuario AND c.activo = 1
        AND {norm_sql} = m.nombre_norm
) t GROUP BY accion;''',
        '',
        '-- UPDATE masivo (ejecutar solo tras revisar preview)',
        f'''UPDATE clientes c
INNER JOIN usuarios u ON u.id_usuario = c.id_usuario_FK AND u.activo = 1
INNER JOIN tmp_mayoristas_excel m ON {norm_sql} = m.nombre_norm
SET c.descuento_porcentaje = m.descuento_porcentaje
WHERE c.activo = 1
  AND (c.descuento_porcentaje IS NULL OR CAST(c.descuento_porcentaje AS DECIMAL(5,2)) <> m.descuento_porcentaje);''',
        '',
        '-- NO ENCONTRADOS EN BD',
        f'''SELECT m.hoja, m.nombre_completo, m.descuento_porcentaje
FROM tmp_mayoristas_excel m
WHERE NOT EXISTS (
    SELECT 1 FROM clientes c
    INNER JOIN usuarios u ON u.id_usuario = c.id_usuario_FK AND u.activo = 1
    WHERE c.activo = 1 AND {norm_sql} = m.nombre_norm
)
ORDER BY m.nombre_completo;''',
        '',
        '-- DUPLICADOS EXACTOS EN EXCEL (misma clave normalizada)',
    ])

    for a, b in dups:
        lines.append(
            f"-- {a['nombre_completo']} ({a['descuento_pct']}%) vs {b['nombre_completo']} ({b['descuento_pct']}%) hoja {b['sheet']}"
        )

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text('\n'.join(lines), encoding='utf-8')
    print('SQL:', out)


if __name__ == '__main__':
    main()
