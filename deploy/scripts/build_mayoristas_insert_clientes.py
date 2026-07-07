#!/usr/bin/env python3
"""Genera SQL INSERT directo (sin stored procedure) para clientes mayoristas."""
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

STG = 'stg_mayoristas_descuento'
PWD_HASH = '$2y$12$Q.mDU56uydGodpkS4A.94uqN9cv3caBkYt3a06XrN9AgbqRTeqR8i'
CORREO_EXPR = "CONCAT('mayorista.', s.id, '@mayoristas.local')"
TELEFONO_EXPR = "LEFT(CONCAT('7', LPAD(s.id, 9, '0')), 15)"


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip())


def norm_key(s):
    s = norm(s).lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^a-z0-9 ]', '', s)
    return re.sub(r'\s+', ' ', s).strip()


def split_name(full):
    parts = [p for p in re.split(r'\s+', norm(full)) if p]
    if not parts:
        return 'Cliente', '.', ''
    if len(parts) == 1:
        return parts[0][:50], '.', ''
    if len(parts) == 2:
        return parts[0][:50], parts[1][:25], ''
    return ' '.join(parts[:-2])[:50], parts[-2][:25], parts[-1][:25]


def parse_discount(val):
    if val is None or str(val).strip() in ('', '/', '-'):
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return None
    pct = round(f * 100, 2) if 0 < f <= 1 else round(f, 2)
    return pct if 0 <= pct <= 100 else None


def esc(s):
    return str(s).replace('\\', '\\\\').replace("'", "''")


def norm_name_sql(alias):
    return f"""LOWER(TRIM(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
    CONCAT_WS(' ', {alias}.nombre, {alias}.primer_apellido, IFNULL({alias}.segundo_apellido, '')),
    CHAR(225),'a'),CHAR(233),'e'),CHAR(237),'i'),CHAR(243),'o'),CHAR(250),'u')))"""


def load_rows(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    rows = []
    seen = set()
    for sheet in wb.sheetnames:
        if sheet == 'Hoja1':
            continue
        for r in wb[sheet].iter_rows(min_row=1, values_only=True):
            if not r:
                continue
            name = norm(r[0]) if len(r) > 0 else ''
            if not name or name in ('/', '-'):
                continue
            disc_raw = r[3] if sheet == 'A' and len(r) >= 4 and r[3] is not None else (r[1] if len(r) >= 2 else None)
            pct = parse_discount(disc_raw)
            if pct is None:
                continue
            key = norm_key(name)
            if key in seen:
                continue
            seen.add(key)
            p_n, p_a1, p_a2 = split_name(name)
            rows.append({
                'sheet': sheet,
                'nombre_completo': name,
                'descuento_pct': pct,
                'nombre_norm': key,
                'p_nombre': p_n,
                'p_ap1': p_a1,
                'p_ap2': p_a2,
            })
    return rows


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads' / 'MAYORISTAS-1.xlsx'
    out = Path(__file__).resolve().parents[2] / 'sql' / '2026_07_06_mayoristas_insert_clientes.sql'
    rows = load_rows(xlsx)
    nn = norm_name_sql('u')

    lines = [
        '-- INSERT clientes mayoristas v3 — INSERT directo (sin stored procedure)',
        f'-- Registros unicos Excel: {len(rows)} | Password: Mayorista2026!',
        '-- Ejecutar TODO (Ctrl+A). Si usas transaccion manual, haz COMMIT al final.',
        '',
        'SET NAMES utf8mb4;',
        'SET collation_connection = utf8mb4_unicode_ci;',
        "SET @current_user_id = COALESCE((SELECT MIN(id_usuario) FROM usuarios WHERE activo = 1), 1);",
        '',
        f'DROP TABLE IF EXISTS {STG};',
        f'''CREATE TABLE {STG} (
  id INT AUTO_INCREMENT PRIMARY KEY,
  hoja VARCHAR(10) NOT NULL,
  nombre_completo VARCHAR(255) NOT NULL,
  descuento_porcentaje DECIMAL(5,2) NOT NULL,
  nombre_norm VARCHAR(255) NOT NULL,
  p_nombre VARCHAR(50) NOT NULL,
  p_ap1 VARCHAR(25) NOT NULL,
  p_ap2 VARCHAR(25) NOT NULL DEFAULT '',
  UNIQUE KEY uq_nombre_norm (nombre_norm)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;''',
        '',
    ]

    batch = [
        "('{hs}', '{nc}', {dp:.2f}, '{nn}', '{pn}', '{pa1}', '{pa2}')".format(
            hs=esc(r['sheet']), nc=esc(r['nombre_completo']), dp=r['descuento_pct'],
            nn=esc(r['nombre_norm']), pn=esc(r['p_nombre']), pa1=esc(r['p_ap1']), pa2=esc(r['p_ap2']),
        )
        for r in rows
    ]
    cols = f'INSERT INTO {STG} (hoja, nombre_completo, descuento_porcentaje, nombre_norm, p_nombre, p_ap1, p_ap2) VALUES'
    for i in range(0, len(batch), 50):
        lines.append(cols)
        lines.append(',\n'.join(batch[i:i + 50]) + ';')
        lines.append('')

    not_in_bd = f"""NOT EXISTS (
    SELECT 1 FROM clientes c
    INNER JOIN usuarios u ON u.id_usuario = c.id_usuario_FK AND u.activo = 1
    WHERE c.activo = 1 AND {nn} = s.nombre_norm
)"""

    lines.extend([
        f'SELECT COUNT(*) AS registros_staging FROM {STG};',
        '',
        '-- 1) Usuarios nuevos (solo si no hay cliente con ese nombre)',
        f'''INSERT INTO usuarios (
    nombre, primer_apellido, segundo_apellido, contrasena, correo, telefono,
    id_direccion_FK, activo, correo_verificado_en
)
SELECT
    s.p_nombre,
    s.p_ap1,
    NULLIF(s.p_ap2, ''),
    '{PWD_HASH}',
    {CORREO_EXPR},
    {TELEFONO_EXPR},
    NULL,
    1,
    NULL
FROM {STG} s
WHERE {not_in_bd}
  AND NOT EXISTS (SELECT 1 FROM usuarios ux WHERE LOWER(TRIM(ux.correo)) = LOWER({CORREO_EXPR}))
  AND NOT EXISTS (SELECT 1 FROM usuarios ux WHERE ux.telefono = {TELEFONO_EXPR});''',
        '',
        'SELECT COUNT(*) AS usuarios_despues_paso1 FROM usuarios WHERE activo = 1;',
        '',
        '-- 2) Clientes ligados a esos usuarios por correo generado',
        f'''INSERT INTO clientes (id_usuario_FK, descuento_porcentaje, uso_cfdi, activo)
SELECT u.id_usuario, s.descuento_porcentaje, 'G03', 1
FROM {STG} s
INNER JOIN usuarios u ON LOWER(TRIM(u.correo)) = LOWER({CORREO_EXPR}) AND u.activo = 1
WHERE NOT EXISTS (SELECT 1 FROM clientes c WHERE c.id_usuario_FK = u.id_usuario AND c.activo = 1);''',
        '',
        '-- 3) Actualizar descuento si el cliente ya existia por nombre',
        f'''UPDATE clientes c
INNER JOIN usuarios u ON u.id_usuario = c.id_usuario_FK AND u.activo = 1
INNER JOIN {STG} s ON {nn} = s.nombre_norm
SET c.descuento_porcentaje = s.descuento_porcentaje
WHERE c.activo = 1
  AND (c.descuento_porcentaje IS NULL OR CAST(c.descuento_porcentaje AS DECIMAL(5,2)) <> s.descuento_porcentaje);''',
        '',
        '-- 4) Rol cliente (si existe en roles)',
        f'''INSERT IGNORE INTO usuario_rol (id_usuario_FK, id_rol_FK)
SELECT u.id_usuario, r.id_rol
FROM {STG} s
INNER JOIN usuarios u ON LOWER(TRIM(u.correo)) = LOWER({CORREO_EXPR}) AND u.activo = 1
INNER JOIN roles r ON r.activo = 1 AND LOWER(TRIM(r.nombre_rol)) IN ('cliente', 'clientes')
WHERE NOT EXISTS (
    SELECT 1 FROM usuario_rol ur WHERE ur.id_usuario_FK = u.id_usuario AND ur.id_rol_FK = r.id_rol
);''',
        '',
        'SELECT COUNT(*) AS total_clientes FROM clientes WHERE activo = 1;',
        '''SELECT CAST(descuento_porcentaje AS DECIMAL(5,2)) AS descuento, COUNT(*) AS total
FROM clientes WHERE activo = 1 AND descuento_porcentaje IS NOT NULL
GROUP BY CAST(descuento_porcentaje AS DECIMAL(5,2)) ORDER BY descuento;''',
        '',
        '-- Si total_clientes sigue en 1, revisar errores arriba en la consola (telefono/correo duplicado).',
        f'-- DROP TABLE {STG};',
    ])

    out.write_text('\n'.join(lines), encoding='utf-8')
    print(f'OK {len(rows)} -> {out}')


if __name__ == '__main__':
    main()
